# copy_sheet, copy_sheet_attributes, and copy_cells methods came from StackOverflow as I looked for the best way
# to transfer a worksheet in one book to another book | Source: https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl

from openpyxl import Workbook, load_workbook
import os
from copy import copy

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cell values and styles
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

# Load the destination workbook where we want all of the sheets to end up
master_location = 'source_workbook.xlsx'
master_wb = load_workbook('source_workbook.xlsx')

# Loop through each .xlsx file in our scorecard_package folder, adding each spreadsheet to the destination workbook
directory = 'scorecard_package/'
print("Looping through files in: ", directory, '\n', os.listdir(directory))

for file in os.listdir(directory): # Access each file in a directory
    if file.endswith('.xlsx'): # Access only the .xlsx files
        
        def buildSheetName(firstinitial, lastinitial):
            s =  ''.join([firstinitial, lastinitial, '-scorecard'])
            #print("Name created for the new sheet: ", s)
            return s

        target_wb = load_workbook(f'{os.path.join(directory, file)}')
        target_sheet = target_wb['scorecard']

        name = target_sheet['a1'].value # Capture name of the submitter from cell a1
        splt_name = name.lower().split(' ')
        
        new_master_sheet = master_wb.create_sheet(f'{buildSheetName( splt_name[0][0], splt_name[1][0] ) }' )

        copy_sheet(new_master_sheet, target_sheet)

        print(f'Scorecard submitted for: {name}')

        master_wb.save('newlyCreatedBook.xlsx')

print(f'All scorecards in {directory} have been added to {master_location}!\n')


