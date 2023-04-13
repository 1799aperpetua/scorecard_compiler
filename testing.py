from openpyxl import load_workbook

from CopySheet import CopySheet # will call CopySheet.copy_sheet(source_sheet, target_sheet)
'''
Source:  Sheet containing the data (submitted scorecard)
Target:  (Blank) Sheet that we are copying the source sheet into
'''

target_wb = load_workbook("source_workbook.xlsx")
new_target_ws = target_wb.create_sheet('NewSheet')
target_wb.save("NewBook.xlsx")

source_wb = load_workbook("source_scorecard.xlsx")

source_ws = source_wb['MASTER']

print("Scorecard A1:", source_ws['c2'].value)
print("Blank Target Sheet:", new_target_ws['a1'].value)

CopySheet.copy_sheet(source_ws, new_target_ws)
target_wb.save("NewBook.xlsx")
