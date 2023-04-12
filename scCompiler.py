import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog
import os
from CopySheet import CopySheet
from openpyxl import load_workbook

class App(ctk.CTk):

    def __init__(self):
        super().__init__()

        self.geometry("1000x500")
        self.title("Scorecard Compiler")
        self.minsize(900, 400)

        self.open_file_bool = False
        self.open_dir_bool = False

        # Entry for the location of name in the scorecard template
        self.choose_name_cell_label = ctk.CTkLabel(master = self, text = "What cell is your submitter name located in?")
        self.choose_name_cell_label.pack(padx = 10, pady = 5)
        self.submitter_name_var = tk.StringVar(value = "C2")
        self.submitter_name_location_entry = ctk.CTkEntry(master = self, textvariable = self.submitter_name_var, placeholder_text = "Input Cell")
        self.submitter_name_location_entry.pack(padx = 10, pady = 5)

        # Entry for name of the scorecard sheets
        self.choose_sheet_name_label = ctk.CTkLabel(master = self, text = "What is the name of the scorecard sheets?")
        self.choose_sheet_name_label.pack(padx = 10, pady = 5)
        self.scorecard_name_var = tk.StringVar(value = "MASTER")
        self.sheet_name_entry = ctk.CTkEntry(master = self, textvariable = self.scorecard_name_var, placeholder_text = "Input Sheet Name")
        self.sheet_name_entry.pack(padx = 10, pady = 5)

        # Label for our buttons
        self.buttons_label = ctk.CTkLabel(master = self, text = "Use the buttons below to choose a file that you want a package of scorecard sheets added to!")
        self.buttons_label.pack(padx = 10, pady = 15)

        # select master file button
        self.master_file_button = ctk.CTkButton(master = self, text = "What file do you want to add your package to?", command = self.openFile)
        self.master_file_button.pack(padx = 10, pady = 10)

        # select package button
        self.package_location_button = ctk.CTkButton(master = self, text = "What folder of scorecards do you want to add?", command = self.openDirectory)
        self.package_location_button.pack(padx = 10, pady = 10)
        
        # execute button
        self.execute_button = ctk.CTkButton(master = self, text = "Compile Scorecards", command = self.transferScorecards)
        self.execute_button.pack(padx = 10, pady = 15)

        # WOULD LOVE to determine how to change the color of these buttons depending on whether or not they've been run/are ready to run
        

    def captureSubmitterName(self):
        print("Submitter Name Entry", self.submitter_name_location_entry.get())
        return self.submitter_name_location_entry.get()

    def captureSheetName(self):
        print("Sheet Name Entry: ", self.sheet_name_entry.get())
        return self.sheet_name_entry.get()

    def openFile(self):
        
        file = filedialog.askopenfile(mode = 'r', filetypes = [("XLSX Files", "*.xlsx")])

        def captureFilename(string):
            '''
            Helper function that strips the location of the file off and just returns the filename
            '''
            i = -1
            ltrs = []
            while string[i] != '/': # add the letters of our file name (in reverse) to a list and stop when we reach the slash indicating that we've reached a folder
                ltrs.append(string[i])
                i -= 1
            
            s = []
            for i in range(len(ltrs)):
                s.append(ltrs.pop())

            return "".join(s)

        if file:
            filepath = os.path.abspath(file.name)
            self.current_master_file = str(filepath)
            master_file_full_name = str(filepath)
            master_file_name = captureFilename(master_file_full_name)
            self.master_file_text = ctk.CTkLabel(master = self, text = "Selected File: " + master_file_name).pack(padx = 10, pady = 5)
            self.open_file_bool = True
            #self.master_file_text.text = str(filepath)

    def openDirectory(self):
        dir = filedialog.askdirectory()
        self.current_chosen_dir = str(dir)
        self.package_location_text = ctk.CTkLabel(master = self, text = "Package Location: \n" + str(dir)).pack(padx = 10, pady = 5)
        self.open_dir_bool = True
        #self.package_location_text.text = str(dir)

    def transferScorecards(self):
        try:
            if self.open_dir_bool is True and self.open_file_bool is True:
            
                master_wb = load_workbook(str(self.current_master_file))

                for file in os.listdir(str(self.current_chosen_dir)): # Access each file in a directory
                    if file.endswith('.xlsx'): # Access only the .xlsx files
                        
                        def buildSheetName(firstinitial, lastinitial):
                            '''
                            Helper function to build the name of a new sheet
                            '''
                            s =  ''.join([firstinitial, lastinitial, '-scorecard'])
                            return s

                        # Load the workbook we're accessing and access the scorecard sheet
                        print("Current Selected Directory: ", self.current_chosen_dir)
                        print("Current File:", file)
                        print("=====================================")
                        target_wb = load_workbook(f'{os.path.join(str(self.current_chosen_dir), file)}')
                        sheet_name = self.captureSheetName()
                        target_sheet = target_wb[f'{sheet_name}']

                        # User's name - Used to create a new sheet with their initials
                        name_loc = self.captureSubmitterName()
                        name = target_sheet[f'{name_loc}'].value 
                        splt_name = name.lower().split(' ')
                        sheet_name = buildSheetName(splt_name[0][0], splt_name[1][0])

                        # Create a new sheet in our master workbook titled "submitterInitials-scorecard"
                        new_master_sheet = master_wb.create_sheet(f'{sheet_name}')

                        # Transfer the currently accessed sheet, into the newly created "submitterInitials-scorecard" sheet
                        CopySheet.copy_sheet(new_master_sheet, target_sheet)

                        print(f'Scorecard submitted for: {name}')

                        new_filename = 'newlyCreatedBook.xlsx'

                        master_wb.save(new_filename)

                        self.master_success_text = ctk.CTkLabel(master = self, text = f"Success! Scorecard sheet added to {new_filename}").pack(padx = 10, pady = 5)
                    else:
                        self.master_error_text = ctk.CTkLabel(master = self, text = "Looks like you haven't selected a source file, a package, or either").pack(padx = 10, pady = 5)
        except:
            self.master_error_text = ctk.CTkLabel(master = self, text = "Error at some point in processing").pack(padx = 10, pady = 5)
                

if __name__ == "__main__":
    app = App()
    app.mainloop()

# Question:  On our scorecard, there's an aggregator at the bottom compiling the scores within each person's scorecard.  Do we need a script to compile all of the worksheets 
# in our master book?